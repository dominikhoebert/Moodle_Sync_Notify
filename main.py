import os
import json

import pandas as pd
import openpyxl
from exchangelib import DELEGATE, Account, Credentials, Message, HTMLBody
import markdown
from markdown.extensions.tables import TableExtension

from dataclasses import dataclass

filename = "/Users/dominik/Downloads/SYTxHIT-WiSe-2223.xlsx"
#filename = "C:/Users/Dominik/Dropbox/TGM/Systemtechnik SYT/Schuljahr 22_23/SYTxHIT-WiSe-2223.xlsx"
#subject = "[INSY] Aktueller Notenstand"

templates_folder = "templates"
email_column = "Email"


def filter_df(df):
    #df = df[df["Negative Kompetenzen"].notnull()]
    #df = df[df["Schüler"] == "D1"]
    df = df[df["Klasse"] == "3DHIT"]
    #df = df[df["Klasse"].notnull()]
    return df


def main():
    import warnings
    warnings.simplefilter("ignore")

    with open("data/credentials.json", "r") as j:
        credentials = json.load(j)

    sender_email = credentials["email"]
    password = credentials["password"]

    file = openpyxl.load_workbook(filename, data_only=True)

    with open("replacements.json", "r", encoding="ISO-8859-1") as j:
        replacements = json.load(j)

    for i, sheet in enumerate(file.sheetnames):
        print(f"[{i + 1}] {sheet.title()}")

    try:
        choice = int(input("Choose a sheet: ")) - 1
    except ValueError:
        print("Invalid input")
        exit()

    if choice < 0 or choice > len(file.sheetnames):
        print("Invalid input")
        exit()

    ws = file[file.sheetnames[choice]]
    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)

    templates = [f for f in os.listdir(templates_folder)]

    for file in templates:
        print(f"[{templates.index(file) + 1}] {file}")

    try:
        choice = int(input("Choose a Template: ")) - 1
    except ValueError:
        print("Invalid input")
        exit()

    if choice < 0 or choice > len(templates):
        print("Invalid input")
        exit()

    with open(f"{templates_folder}/{templates[choice]}", "r") as f:
        template = f.read()

    emails = []

    df = filter_df(df)

    if template.startswith("Subject:"):
        lines = template.splitlines()
        subject = lines[0][8:].strip()
        template = "\n".join(lines[1:])

    for i, row in df.iterrows():
        message = template
        if templates[choice].endswith(".md"):
            message = markdown.markdown(message, extensions=[TableExtension()])
        for column in columns:
            part = str(row[column])
            if part in replacements:
                part = replacements[part]
            message = message.replace(f"[{column}]", part)
        emails.append(Email(row[email_column], subject, message))

    r = 3
    if len(emails) < 3:
        r = len(emails)
    for i in range(r):
        print(emails[i].adress, "\n", emails[i].subject, "\n", emails[i].message)

    if input(f"\n\nSend {len(emails)} emails? (y/n)") == "y":
        credentials = Credentials(username=sender_email, password=password)
        exchange_account = Account(
            primary_smtp_address=sender_email, credentials=credentials,
            autodiscover=True, access_type=DELEGATE
        )

        message_ids = []
        for email in emails:
            message = Message(
                account=exchange_account,
                folder=exchange_account.drafts,
                subject=email.subject,
                body=HTMLBody(email.message),
                to_recipients=[email.adress]
            ).save()
            message_ids.append((message.id, message.changekey))

        result = exchange_account.bulk_send(ids=message_ids)
        print(f"\n\n{result.count(True)} emails sent sucessfully.\n{result.count(False)} emails failed.")


@dataclass
class Email:
    adress: str
    subject: str
    message: str


if __name__ == "__main__":
    main()
